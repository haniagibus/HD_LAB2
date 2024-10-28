# -*- encoding: utf-8 -*-
import shutil
from datetime import datetime, timedelta, date
import openpyxl
from faker import Faker
import random
import string
from faker.providers import DynamicProvider
import pandas as pd
import calendar

positions_provider = DynamicProvider(
    provider_name="position",
    elements=["recepcja", "lekarz", "personel medyczny", "pielęgniarka", "lekarz - stażysta"]
)

shifts_provider = DynamicProvider(
    provider_name="shift",
    elements=["7:00-14:00", "10:00-17:00", "14:00-20:00", ""]
)

fake = Faker("pl_PL")
fake.add_provider(positions_provider)
fake.add_provider(shifts_provider)

t1_start = datetime.strptime('2018-01-01', '%Y-%m-%d').date()
t1_end = datetime.strptime('2022-01-01', '%Y-%m-%d').date()
t2_start = datetime.strptime('2022-01-02', '%Y-%m-%d').date()
t2_end = datetime.strptime('2024-01-01', '%Y-%m-%d').date()
def generate_unique_identifier(existing_ids, length):
    """Generuje unikalny identyfikator o podanej długości."""
    while True:
        identifier = ''.join(random.choices(string.ascii_letters + string.digits, k=length))
        if identifier not in existing_ids:  # Sprawdzenie, czy identyfikator jest unikalny
            existing_ids.add(identifier)
            return identifier

# time=1 --> t1
# time=2 --> t2

def create_data(time):
    workbook = openpyxl.Workbook()

    #arkusz 1
    if time==1:
        sheet = workbook.active
        sheet.title = "Pracownicy"
        headers = ["Identyfikator", "PESEL", "Imie", "Nazwisko", "Data urodzenia", "Stanowisko", "Data zatrudnienia",
                    "Data zwolnienia"]
        sheet.append(headers)
        existing_ids = set()
    else:
        kopia_pliku = "excel_dyrektora_merged.xlsx"
        shutil.copy("excel_dyrektora.xlsx", kopia_pliku)
        # Otwórz skopiowany plik do edycji
        workbook = openpyxl.load_workbook(kopia_pliku)
        sheet = workbook.active
        workbook_orig = openpyxl.load_workbook("excel_dyrektora.xlsx")
        sheet_orig = workbook_orig.active
        existing_ids = [sheet_orig.cell(row=i, column=1).value for i in range(2, sheet_orig.max_row + 1)]

    if time==1:
        for _ in range(1000):
            ID = generate_unique_identifier(existing_ids, 12)
            PESEL = fake.unique.random_int(10000000000, 99999999999)
            imie = fake.first_name()
            nazwisko = fake.last_name()
            data_urodzenia = fake.date_between(start_date=datetime.strptime('1955-01-01', '%Y-%m-%d').date(),end_date=datetime.strptime('1997-01-01', '%Y-%m-%d').date())
            stanowisko = fake.position()
            data_zatrudnienia = fake.date_between(data_urodzenia + timedelta(days=8760),end_date=t1_end)
            if random.choices([True, False], weights=[0.7, 0.3])[0]:  # 70% True, 30% False
                data_zwolnienia = None
            else:
                data_zwolnienia = fake.date_between(start_date=data_zatrudnienia, end_date=datetime.strptime('2024-01-01', '%Y-%m-%d'))
            row = [ID, PESEL, imie, nazwisko, data_urodzenia, stanowisko, data_zatrudnienia, data_zwolnienia]
            sheet.append(row)

        workbook.save("excel_dyrektora.xlsx")

        #arkusz 2
        pracownicy_ids = pd.read_excel("excel_dyrektora.xlsx", usecols="A")
        pracownicy_names = pd.read_excel("excel_dyrektora.xlsx", usecols="B")
        pracownicy_surnames = pd.read_excel("excel_dyrektora.xlsx", usecols="C")
        pracownicy_positions = pd.read_excel("excel_dyrektora.xlsx", usecols="D")

        sheet = workbook.create_sheet("harmonogram")

        headers = ["Identyfikator", "Imie", "Nazwisko", "Stanowisko"]
        num_days = calendar.monthrange(date.today().year, date.today().month)[1]
        for day in range(1, num_days):
            headers.append(str(day) + "." + str(date.today().month))

        sheet.append(headers)

        for i in range(1000):
            ID = pracownicy_ids.iloc[i, 0]
            imie = pracownicy_names.iloc[i, 0]
            nazwisko = pracownicy_surnames.iloc[i, 0]
            stanowisko = pracownicy_positions.iloc[i, 0]

            row = [ID, imie, nazwisko, stanowisko]
            for day in range(1, num_days):
                row.append(fake.shift())

            sheet.append(row)

        workbook.save("excel_dyrektora.xlsx")
    elif time==2:
        for row in range(2, sheet.max_row + 1):
            stanowisko_komorka = sheet.cell(row=row, column=6)
            if stanowisko_komorka.value == "lekarz - stażysta":
                # 50% szansy na zmianę na "lekarz"
                if random.random() < 0.5:
                    stanowisko_komorka.value = "lekarz"

            stanowisko_komorka = sheet.cell(row=row, column=8)
            if stanowisko_komorka.value == None:
                # 95% szansy na zwolnienie
                if random.random() < 0.05:
                    stanowisko_komorka.value = fake.date_time_between(t2_start,t2_end)

        workbook.save(kopia_pliku)

create_data(1)
create_data(2)