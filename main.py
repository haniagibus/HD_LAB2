import os
import shutil
from faker import Faker
from faker.providers import DynamicProvider
from datetime import datetime, timedelta
import csv
import random
import string
import openpyxl
import pandas as pd

specjalization_provider = DynamicProvider(
    provider_name="medical_specialization",
    elements=["Kardiolog", "Pediatra", "Ginekolog", "Dermatolog", "Neurolog",
              "Ortopeda", "Chirurg", "Psychiatra", "Endokrynolog", "Onkolog",
              "Okulista", "Laryngolog", "Reumatolog",
              "Urolog", "Chirurg plastyczny", "Chirurg naczyniowy", "Medycyna rodzinna",
              "Medycyna pracy", "Medycyna sportowa", "Pneumolog", "Immunolog",
              "Patolog", "Radiolog", "Nefrolog", "Ginekolog-onkolog",
              "Ortopeda dziecięcy", "Medycyna ratunkowa"],
)

fake = Faker("pl_PL")
fake.add_provider(specjalization_provider)

my_sentence_list = [
    'Pacjenta boli brzuch.',
    'Pacjent skarży się na ból głowy.',
    'Pacjent ma gorączkę i dreszcze.',
    'Pacjent odczuwa osłabienie.',
    'Pacjent ma duszności i kaszel.',
    'Pacjent zgłasza katar i ból gardła.',
    'Pacjent ma zawroty głowy.',
    'Pacjent skarży się na wymioty.',
    'Pacjent ma biegunkę.',
    'Pacjent zgłasza ból mięśni i stawów.',
    'Pacjent odczuwa pieczenie oczu.',
    'Pacjent ma problemy z trawieniem.',
    'Pacjent zgłasza opuchliznę nóg.',
    'Pacjent odczuwa swędzenie skóry.',
    'Pacjent ma wysokie ciśnienie krwi.',
    'Pacjent ma problemy ze wzrokiem.',
    'Pacjent odczuwa zmęczenie i senność.',
    'Pacjent ma ból zęba.',
    'Pacjent skarży się na bóle kręgosłupa.',
    'Pacjent zgłasza łzawienie oczu.',
    'Pacjent odczuwa drżenie rąk.',
    'Pacjent ma objawy przeziębienia.',
    'Pacjent skarży się na ból zatok.',
    'Pacjent odczuwa bóle brzucha po jedzeniu.',
    'Pacjent zgłasza trudności w oddychaniu.',
    'Pacjent ma swędzenie i zaczerwienienie skóry.',
    'Pacjent skarży się na szumy uszne.',
    'Pacjent odczuwa silne bóle w klatce piersiowej.',
    'Pacjent ma problem z pamięcią.',
    'Pacjent zgłasza uczucie niepokoju.',
    'Pacjent ma problemy ze snem.',
    'Pacjent skarży się na ból w stawie kolanowym.',
    'Pacjent ma zgagę.',
    'Pacjent odczuwa dreszcze bez gorączki.',
    'Pacjent ma wrażliwość na światło.',
    'Pacjent skarży się na bóle w odcinku szyjnym kręgosłupa.',
    'Pacjent zgłasza bóle głowy przy wysiłku.',
    'Pacjent ma problemy z równowagą.',
    'Pacjent odczuwa sztywność mięśni.',
    'Pacjent ma objawy alergii pokarmowej.',
    'Pacjent zgłasza trudności w przełykaniu.',
    'Pacjent odczuwa ból w klatce piersiowej przy głębokim wdechu.',
    'Pacjent ma mrowienie w kończynach.',
    'Pacjent zgłasza częste bóle głowy związane z migreną.',
    'Pacjent skarży się na kłopoty z koncentracją.',
    'Pacjent ma napady paniki.'
]

badania = [
    "morfologia krwi",
    "tomografia komputerowa",
    "badanie moczu",
    "USG jamy brzusznej",
    "EKG (elektrokardiogram)",
    "EEG (elektroencefalogram)",
    "test na cholesterol",
    "badanie poziomu cukru",
    "badanie wzroku",
    "spirometria",
    "badanie słuchu",
    "próba wysiłkowa",
    "RTG klatki piersiowej",
    "mammografia",
    "test na HIV",
    "badanie poziomu witaminy D",
    "test PSA",
    "badanie TSH (tarczyca)",
    "test na alergie",
    "markery nowotworowe",
    "biopsja skóry",
    "densytometria (gęstość kości)",
    "badanie kału",
    "test wątrobowy",
    "test nerkowy",
    "badanie prostaty",
    "test na anemię",
    "glukoza na czczo",
    "posiew bakteryjny",
    "test na infekcje",
    "badanie hormonów"
]

tableSize=200
mainTableSize=1000

def generate_unique_identifier(existing_ids, length):
    while True:
        identifier = ''.join(random.choices(string.ascii_letters + string.digits, k=length))
        if identifier not in existing_ids:  # Sprawdzenie, czy identyfikator jest unikalny
            existing_ids.add(identifier)
            return identifier

def generate_random_time_within_range(start_hour=7, end_hour=19):
    hour = random.randint(start_hour, end_hour)
    minute = random.choice([0, 10, 20, 30, 40, 50])
    return f"{hour:02d}:{minute:02d}"


def biased_random_int(low, high, bias=0.9):
    if random.random() < bias:
        # generate low number
        return fake.random_int(min=low, max=(low + high) // 8)
    else:
        # generate high number
        return fake.random_int(min=(low + high) // 8 + 1, max=high)


existing_ids = set()
id_lekarzy = []


def generate_data(t):
    if t == 1:
        dir = ".\\t1\\"
        mode = "w"
        time_start = datetime.strptime('2018-01-01', '%Y-%m-%d').date()
        time_end = datetime.strptime('2022-01-01', '%Y-%m-%d').date()
    else:
        dir = ".\\t2\\"
        mode = "a"
        time_start = datetime.strptime('2022-01-02', '%Y-%m-%d').date()
        time_end = datetime.strptime('2024-01-01', '%Y-%m-%d').date()

    excel = openpyxl.load_workbook(dir + "excel_dyrektora.xlsx")
    pracownicy = excel["Pracownicy"]

    # Pacjenci
    with open(dir + "dane_pacjenci.csv", mode=mode, newline="") as pacjenci_csv:
        writer = csv.writer(pacjenci_csv)
        if t == 1:
            writer.writerow(["ID_pacjenta", "Imię", "Nazwisko", "Nr_telefonu",
                             "Miejscowość", "Ulica_i_numer_domu", "Pesel"])

        for i in range(tableSize):
            ID_pacjenta = fake.unique.random_int(100000000, 999999999)
            imie = fake.first_name()
            nazwisko = fake.last_name()
            nr_telefonu = fake.unique.numerify("### ### ###")
            miejscowosc = fake.city()
            ulica_i_numer_domu = fake.street_address()
            pesel = fake.unique.random_int(10000000000, 99999999999)

            writer.writerow([ID_pacjenta, imie, nazwisko, nr_telefonu, miejscowosc, ulica_i_numer_domu, pesel])

    # Lekarze
    with open(dir + "dane_lekarze.csv", mode=mode, newline="") as lekarze_csv:
        writer = csv.writer(lekarze_csv)
        if t == 1:
            writer.writerow(["Identyfikator", "Imię", "Nazwisko", "Specjalizacja"])

        for row in pracownicy.iter_rows(min_row=2, values_only=True):
            identyfikator, pesel, imie, nazwisko, data_urodzenia, stanowisko, data_zatrudnienia, data_zwolnienia = row
            if stanowisko == 'lekarz' and identyfikator not in id_lekarzy:
                id_lekarzy.append(identyfikator)
                specjalizacja = fake.medical_specialization()
                writer.writerow([identyfikator, imie, nazwisko, specjalizacja])

    lekarze_file = pd.read_csv(dir + "dane_lekarze.csv", encoding="windows-1250")
    pacjenci_file = pd.read_csv(dir + "dane_pacjenci.csv", encoding="windows-1250")

    id_wszyscy_pacjenci = pacjenci_file['ID_pacjenta'].tolist()
    id_wszyscy_lekarze = lekarze_file['Identyfikator'].tolist()

    # Badania
    with open(dir + "dane_badania.csv", mode=mode, newline="") as badania_csv:
        writer = csv.writer(badania_csv)
        if t == 1:
            writer.writerow(["nazwa_badania", "czas_trwania", "koszt"])

            for i in range(len(badania)):
                nazwa_badania = badania[i]
                czas_trwania = fake.random_int(5, 300, 5)
                koszt = biased_random_int(0, 10000)

                writer.writerow([nazwa_badania, czas_trwania, koszt])

    # Wizyty
    recepcjonistki = []
    for row in pracownicy.iter_rows(min_row=2, values_only=True):
        identyfikator, pesel, imie, nazwisko, data_urodzenia, stanowisko, data_zatrudnienia, data_zwolnienia = row
        if stanowisko == 'recepcja':
            recepcjonistki.append(identyfikator)

    with open(dir + "dane_wizyty.csv", mode=mode, newline="") as wizyty_csv:
        writer = csv.writer(wizyty_csv)
        if t == 1:
            writer.writerow(["ID_wizyty", "data_umówienia", "data_rozpoczecia", "dolegliwosci", "kwota", "godzina",
                             "czy_odbyta", "ID_recepcjonistki", "ID_pacjenta", "ID_lekarza"])

        for i in range(mainTableSize):
            ID_wizyty = generate_unique_identifier(existing_ids, 10)
            data_umowienia = fake.date_between(time_start, time_end - timedelta(days=180))
            data_rozpoczecia = fake.date_between(start_date=data_umowienia,
                                                 end_date=data_umowienia + timedelta(days=180))
            dolegliwosci = fake.sentence(ext_word_list=my_sentence_list)
            assert len(dolegliwosci) <= 1000
            kwota = round(random.uniform(40, 600), 2)
            godzina = generate_random_time_within_range()
            czy_odbyta = random.choices(["TAK", "NIE"], weights=[0.7, 0.3], k=1)[0]
            ID_recepcjonistki = random.choice(recepcjonistki)
            ID_pacjenta = random.choice(id_wszyscy_pacjenci)
            ID_lekarza = random.choice(id_wszyscy_lekarze)
            writer.writerow(
                [ID_wizyty, data_umowienia, data_rozpoczecia, dolegliwosci, kwota, godzina, czy_odbyta,
                 ID_recepcjonistki,
                 ID_pacjenta, ID_lekarza])

    wizyty_file = pd.read_csv(dir + "dane_wizyty.csv", encoding='windows-1250')
    badania_file = pd.read_csv(dir + "dane_badania.csv", encoding='windows-1250')

    wizyty_ids = wizyty_file['ID_wizyty'].tolist()
    badania_names = badania_file['nazwa_badania'].tolist()

    # Recepty
    with open(dir + "dane_recepty.csv", mode=mode, newline="") as recepty_csv:
        writer = csv.writer(recepty_csv)
        if t == 1:
            writer.writerow(["ID_recepty", "waznosc", "czy_wykupiona", "data_wystawienia", "id_wizyty"])

        for i in range(tableSize):
            ID_recepty = fake.unique.random_int(100000000000000, 999999999999999)
            waznosc = fake.random_int(30, 180, 30)
            czy_wykupiona = fake.boolean()
            data_wystawienia = fake.date_between(time_start, time_end)
            id_wizyty = random.choice(wizyty_ids)

            writer.writerow([ID_recepty, waznosc, czy_wykupiona, data_wystawienia, id_wizyty])

    # Zlecenia
    with open(dir + "dane_zlecen.csv", mode=mode, newline="") as zlecenia_csv:
        writer = csv.writer(zlecenia_csv)
        if t == 1:
            writer.writerow(["ID_wizyty", "nazwa_badania"])

        for i in range(tableSize):
            ID_wizyty = random.choice(wizyty_ids)
            nazwa_badania = random.choice(badania_names)

            writer.writerow([ID_wizyty, nazwa_badania])


generate_data(1)

src_dir = ".\\t1\\"
dst_dir = ".\\t2\\"
os.makedirs(dst_dir, exist_ok=True)

for file in os.listdir(src_dir):
    if file.endswith(".csv"):
        shutil.copy(os.path.join(src_dir, file), os.path.join(dst_dir, file))

generate_data(2)

# Oświadczamy, że treści wygenerowane przy pomocy z GenAI poddałyśmy krytycznej analizie i zweryfikowałyśmy.
# Korzystałyśmy za zgodą prowadzącego z następujących narzędzi o potencjalnie wysokim stopniu ingerencji:
# ChatGPT - wygenerowanie danych dla: my_sentence_list, specjalization_provider
